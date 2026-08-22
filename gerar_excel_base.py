
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime, time, timedelta

# ============================================================
# CONFIGURAÇÕES VISUAIS
# ============================================================

COR_PRIMARIA = "1F4E79"
COR_SECUNDARIA = "2E75B6"
COR_ACENTO = "ED7D31"
COR_SUCESSO = "70AD47"
COR_ALERTA = "FF0000"
COR_FUNDO = "F2F2F2"
COR_BRANCO = "FFFFFF"
COR_CINZA_ESCURO = "404040"
COR_CINZA_CLARO = "D9D9D9"

FONTE_TITULO = Font(name="Calibri", size=18, bold=True, color=COR_BRANCO)
FONTE_SUBTITULO = Font(name="Calibri", size=12, bold=True, color=COR_BRANCO)
FONTE_CABECALHO = Font(name="Calibri", size=10, bold=True, color=COR_BRANCO)
FONTE_NORMAL = Font(name="Calibri", size=10, color="000000")
FONTE_BOTAO = Font(name="Calibri", size=9, bold=True, color=COR_BRANCO)

PREECHIMENTO_PRIMARIO = PatternFill(start_color=COR_PRIMARIA, end_color=COR_PRIMARIA, fill_type="solid")
PREECHIMENTO_SECUNDARIO = PatternFill(start_color=COR_SECUNDARIA, end_color=COR_SECUNDARIA, fill_type="solid")
PREECHIMENTO_ACENTO = PatternFill(start_color=COR_ACENTO, end_color=COR_ACENTO, fill_type="solid")
PREECHIMENTO_SUCESSO = PatternFill(start_color=COR_SUCESSO, end_color=COR_SUCESSO, fill_type="solid")
PREECHIMENTO_ALERTA = PatternFill(start_color=COR_ALERTA, end_color=COR_ALERTA, fill_type="solid")
PREECHIMENTO_FUNDO = PatternFill(start_color=COR_FUNDO, end_color=COR_FUNDO, fill_type="solid")
PREECHIMENTO_CINZA_ESCURO = PatternFill(start_color=COR_CINZA_ESCURO, end_color=COR_CINZA_ESCURO, fill_type="solid")
PREECHIMENTO_CINZA_CLARO = PatternFill(start_color=COR_CINZA_CLARO, end_color=COR_CINZA_CLARO, fill_type="solid")

BORDA_FINA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

BORDA_MEDIA = Border(
    left=Side(style="medium", color=COR_SECUNDARIA),
    right=Side(style="medium", color=COR_SECUNDARIA),
    top=Side(style="medium", color=COR_SECUNDARIA),
    bottom=Side(style="medium", color=COR_SECUNDARIA)
)

ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALINHAMENTO_ESQUERDA = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALINHAMENTO_DIREITA = Alignment(horizontal="right", vertical="center")

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def aplicar_estilo_celula(celula, fonte, preenchimento, alinhamento, borda=None):
    celula.font = fonte
    celula.fill = preenchimento
    celula.alignment = alinhamento
    if borda:
        celula.border = borda

def criar_titulo_aba(ws, titulo, subtitulo=""):
    ws.merge_cells("A1:Z1")
    ws["A1"] = titulo
    ws["A1"].font = FONTE_TITULO
    ws["A1"].fill = PREECHIMENTO_PRIMARIO
    ws["A1"].alignment = ALINHAMENTO_CENTRO
    ws.row_dimensions[1].height = 30

    if subtitulo:
        ws.merge_cells("A2:Z2")
        ws["A2"] = subtitulo
        ws["A2"].font = FONTE_SUBTITULO
        ws["A2"].fill = PREECHIMENTO_SECUNDARIO
        ws["A2"].alignment = ALINHAMENTO_CENTRO
        ws.row_dimensions[2].height = 20

def criar_cabecalho_tabela(ws, linha, colunas, larguras=None):
    for i, coluna in enumerate(colunas, 1):
        celula = ws.cell(row=linha, column=i, value=coluna)
        aplicar_estilo_celula(celula, FONTE_CABECALHO, PREECHIMENTO_SECUNDARIO, ALINHAMENTO_CENTRO, BORDA_FINA)
        if larguras and i <= len(larguras):
            ws.column_dimensions[get_column_letter(i)].width = larguras[i-1]

def ajustar_largura_colunas(ws, colunas, largura_padrao=15):
    for i in range(1, colunas + 1):
        letra = get_column_letter(i)
        if ws.column_dimensions[letra].width < largura_padrao:
            ws.column_dimensions[letra].width = largura_padrao

# ============================================================
# CRIAR ARQUIVO
# ============================================================

wb = openpyxl.Workbook()

# Remover aba padrão
wb.remove(wb.active)

# ============================================================
# ABA 1: DADOS
# ============================================================

ws_dados = wb.create_sheet("DADOS")

criar_titulo_aba(ws_dados, "DADOS - Ordem de Produção", "Cadastro e cálculo das OPs")
ws_dados.row_dimensions[3].height = 25

colunas_dados = [
    "OP", "Produto", "Dosagem", "Quantidade", "Unidade",
    "Caixas", "Máquina", "Sequência", "Data Planejada",
    "Velocidade", "OEE", "Capacidade_h", "Produção_h", "Setup_h",
    "Duração Base_h", "Duração Total (h)", "Início", "Fim",
    "Status", "Atraso_h", "Eventos_h", "Refeição_h", "Início Original", "Fim Original", "Observação"
]

larguras_dados = [12, 20, 12, 12, 10, 10, 15, 10, 16, 10, 8, 12, 12, 10, 14, 16, 16, 16, 14, 10, 10, 12, 16, 16, 25]

criar_cabecalho_tabela(ws_dados, 3, colunas_dados, larguras_dados)
ajustar_largura_colunas(ws_dados, len(colunas_dados), 14)

# Dados de exemplo
dados_exemplo = [
    ["OP001", "PURAN", "25 mcg", 1000, "COMP", 200, "FETTE 2090", 1, datetime(2026, 8, 25, 5, 0), 500, 0.85, 425, 0.47, 0.5, 0.97, 1.0, datetime(2026, 8, 25, 5, 0), datetime(2026, 8, 25, 6, 0), "PLANEJADO", 0, 0, 0, "", "", ""],
    ["OP002", "PURAN", "50 mcg", 800, "COMP", 160, "FETTE 2", 1, datetime(2026, 8, 25, 6, 30), 450, 0.80, 360, 0.44, 0.5, 1.33, 1.33, datetime(2026, 8, 25, 6, 30), datetime(2026, 8, 25, 8, 0), "PLANEJADO", 0, 0, 0, "", "", ""],
    ["OP003", "DIPIRONA", "1g", 500, "CX", 500, "MEDISEAL", 1, datetime(2026, 8, 25, 5, 0), 600, 0.75, 450, 1.11, 0.5, 1.61, 1.61, datetime(2026, 8, 25, 5, 0), datetime(2026, 8, 25, 6, 45), "PLANEJADO", 0, 0, 0, "", "", ""],
]

for i, linha_dados in enumerate(dados_exemplo, 4):
    for j, valor in enumerate(linha_dados, 1):
        celula = ws_dados.cell(row=i, column=j, value=valor)
        aplicar_estilo_celula(celula, FONTE_NORMAL, PREECHIMENTO_FUNDO, ALINHAMENTO_CENTRO if j != 25 else ALINHAMENTO_ESQUERDA, BORDA_FINA)
        if j in [16, 17, 18]:  # Datas
            celula.number_format = "dd/mm/yyyy hh:mm"
        elif j in [10, 11, 12, 13, 14, 15, 20, 21]:  # Números
            celula.number_format = "0.00"

ws_dados.freeze_panes = "A4"

# ============================================================
# ABA 2: PLANEJAMENTO
# ============================================================

ws_plan = wb.create_sheet("PLANEJAMENTO")

criar_titulo_aba(ws_plan, "PLANEJAMENTO DA PRODUÇÃO", "Timeline visual e cards das OPs")
ws_plan.row_dimensions[3].height = 20
ws_plan.row_dimensions[4].height = 25

# Cabeçalho da timeline
ws_plan.cell(row=4, column=1, value="MÁQUINA").font = FONTE_CABECALHO
ws_plan.cell(row=4, column=1).fill = PREECHIMENTO_CINZA_ESCURO
ws_plan.cell(row=4, column=1).alignment = ALINHAMENTO_CENTRO
ws_plan.cell(row=4, column=1).border = BORDA_FINA

hora_inicio = time(5, 0)
for i in range(2, 50):
    hora = (datetime.combine(datetime.today(), hora_inicio) + timedelta(minutes=30 * (i - 2))).time()
    celula = ws_plan.cell(row=4, column=i, value=hora.strftime("%H:%M"))
    celula.font = FONTE_CABECALHO
    celula.fill = PREECHIMENTO_CINZA_ESCURO
    celula.alignment = ALINHAMENTO_CENTRO
    celula.border = BORDA_FINA
    ws_plan.column_dimensions[get_column_letter(i)].width = 9

# Máquinas
maquinas = ["FETTE 2090", "FETTE 2", "MEDISEAL", "BLISTERFLEX"]
for i, maquina in enumerate(maquinas, 5):
    celula = ws_plan.cell(row=i, column=1, value=maquina)
    celula.font = Font(name="Calibri", size=10, bold=True, color=COR_BRANCO)
    celula.fill = PREECHIMENTO_SECUNDARIO
    celula.alignment = ALINHAMENTO_CENTRO
    celula.border = BORDA_FINA
    ws_plan.row_dimensions[i].height = 60

ws_plan.column_dimensions["A"].width = 20
ws_plan.freeze_panes = "B5"

# Área para botões (linha 1)
ws_plan.row_dimensions[1].height = 35
ws_plan.row_dimensions[2].height = 10

# ============================================================
# ABA 3: RESUMO (DASHBOARD)
# ============================================================

ws_resumo = wb.create_sheet("RESUMO")

criar_titulo_aba(ws_resumo, "DASHBOARD - APS PURAN", "Indicadores e controle da produção")
ws_resumo.row_dimensions[3].height = 25

# KPIs
kpis = [
    ("TOTAL DE OPs", "=COUNTA(DADOS!A:A)-1", "B4"),
    ("CONCLUÍDAS", '=COUNTIF(DADOS!S:S,"CONCLUÍDO")', "E4"),
    ("EM PRODUÇÃO", '=COUNTIF(DADOS!S:S,"EM PRODUÇÃO")', "H4"),
    ("ATRASADAS", '=COUNTIF(DADOS!S:S,"ATRASADO")', "K4"),
    ("TOTAL DE CAIXAS", "=SUM(DADOS!F:F)", "B9"),
    ("HORAS PLANEJADAS", "=SUM(DADOS!P:P)", "E9"),
    ("HORAS DE ATRASO", "=SUM(DADOS!T:T)", "H9"),
    ("OPs COM ATRASO", '=COUNTIF(DADOS!T:T,">0")', "K9"),
]

# Título dos KPIs
ws_resumo["A4"] = "STATUS DAS OPs"
ws_resumo["A4"].font = Font(name="Calibri", size=11, bold=True, color=COR_BRANCO)
ws_resumo["A4"].fill = PREECHIMENTO_SECUNDARIO
ws_resumo["A4"].alignment = ALINHAMENTO_CENTRO
ws_resumo.merge_cells("A4:C4")
ws_resumo.row_dimensions[4].height = 22

for i, (titulo, _, celula_ref) in enumerate(kpis[:4]):
    col = i + 1
    celula_titulo = ws_resumo.cell(row=5, column=col, value=titulo)
    celula_titulo.font = FONTE_CABECALHO
    celula_titulo.fill = PREECHIMENTO_CINZA_CLARO
    celula_titulo.alignment = ALINHAMENTO_CENTRO
    celula_titulo.border = BORDA_FINA
    
    celula_valor = ws_resumo[celula_ref]
    celula_valor.font = Font(name="Calibri", size=16, bold=True, color=COR_PRIMARIA)
    celula_valor.alignment = ALINHAMENTO_CENTRO
    celula_valor.border = BORDA_FINA
    ws_resumo.row_dimensions[6].height = 30

ws_resumo["A9"] = "INDICADORES DE PRODUÇÃO"
ws_resumo["A9"].font = Font(name="Calibri", size=11, bold=True, color=COR_BRANCO)
ws_resumo["A9"].fill = PREECHIMENTO_SECUNDARIO
ws_resumo["A9"].alignment = ALINHAMENTO_CENTRO
ws_resumo.merge_cells("A9:C9")
ws_resumo.row_dimensions[9].height = 22

for i, (titulo, _, celula_ref) in enumerate(kpis[4:]):
    col = i + 1
    celula_titulo = ws_resumo.cell(row=10, column=col, value=titulo)
    celula_titulo.font = FONTE_CABECALHO
    celula_titulo.fill = PREECHIMENTO_CINZA_CLARO
    celula_titulo.alignment = ALINHAMENTO_CENTRO
    celula_titulo.border = BORDA_FINA
    
    celula_valor = ws_resumo[celula_ref]
    celula_valor.font = Font(name="Calibri", size=16, bold=True, color=COR_PRIMARIA)
    celula_valor.alignment = ALINHAMENTO_CENTRO
    celula_valor.border = BORDA_FINA
    ws_resumo.row_dimensions[11].height = 30

ajustar_largura_colunas(ws_resumo, 15, 18)
ws_resumo.freeze_panes = "A13"

# ============================================================
# ABA 4: RECURSOS
# ============================================================

ws_rec = wb.create_sheet("RECURSOS")

criar_titulo_aba(ws_rec, "RECURSOS - Cadastro de Máquinas", "Parâmetros produtivos por recurso")
ws_rec.row_dimensions[3].height = 25

colunas_rec = ["Máquina", "Velocidade", "Unidade Velocidade", "OEE", "Comprimidos por caixa", "Setup padrão", "Disponibilidade", "Observação"]
larguras_rec = [20, 12, 16, 10, 20, 14, 14, 30]

criar_cabecalho_tabela(ws_rec, 3, colunas_rec, larguras_rec)

recursos_exemplo = [
    ["FETTE 2090", 500, "cx/h", 0.85, 20, 0.5, 0.90, "Envelope"],
    ["FETTE 2", 450, "cx/h", 0.80, 20, 0.5, 0.90, "Envelope"],
    ["MEDISEAL", 600, "cx/h", 0.75, 20, 0.5, 0.85, "Blister"],
    ["BLISTERFLEX", 550, "cx/h", 0.80, 20, 0.5, 0.85, "Blister"],
]

for i, linha_dados in enumerate(recursos_exemplo, 4):
    for j, valor in enumerate(linha_dados, 1):
        celula = ws_rec.cell(row=i, column=j, value=valor)
        aplicar_estilo_celula(celula, FONTE_NORMAL, PREECHIMENTO_FUNDO, ALINHAMENTO_CENTRO, BORDA_FINA)
        if j == 4:  # OEE
            celula.number_format = "0%"
        elif j in [2, 3, 5, 6, 7]:
            celula.number_format = "0.00"

ws_rec.freeze_panes = "A4"

# ============================================================
# ABA 5: EVENTOS
# ============================================================

ws_eventos = wb.create_sheet("EVENTOS")

criar_titulo_aba(ws_eventos, "EVENTOS - Eventos Adicionais", "Registro de eventos que impactam a produção")
ws_eventos.row_dimensions[3].height = 25

colunas_eventos = ["OP", "Tipo", "Duração_h", "Ativo", "Aplicado"]
larguras_eventos = [15, 25, 12, 10, 12]

criar_cabecalho_tabela(ws_eventos, 3, colunas_eventos, larguras_eventos)

eventos_exemplo = [
    ["OP001", "MANUTENÇÃO", 1.0, "SIM", ""],
    ["OP002", "SETUP ADICIONAL", 0.5, "SIM", ""],
]

for i, linha_dados in enumerate(eventos_exemplo, 4):
    for j, valor in enumerate(linha_dados, 1):
        celula = ws_eventos.cell(row=i, column=j, value=valor)
        aplicar_estilo_celula(celula, FONTE_NORMAL, PREECHIMENTO_FUNDO, ALINHAMENTO_CENTRO, BORDA_FINA)
        if j == 3:
            celula.number_format = "0.00"

ws_eventos.freeze_panes = "A4"

# ============================================================
# ABA 6: REFEICAO
# ============================================================

ws_ref = wb.create_sheet("REFEICAO")

criar_titulo_aba(ws_ref, "REFEIÇÃO - Intervalos Fixos", "Configuração dos horários de refeição")
ws_ref.row_dimensions[3].height = 25

colunas_ref = ["ID", "Máquina", "Início", "Duração_h", "Fim", "Ativo", "Observação"]
larguras_ref = [8, 20, 18, 12, 18, 10, 25]

criar_cabecalho_tabela(ws_ref, 3, colunas_ref, larguras_ref)

refeicao_exemplo = [1, "FETTE 2090", datetime(2026, 8, 25, 12, 0), 1.0, datetime(2026, 8, 25, 13, 0), "SIM", "Refeição almoço"]

for j, valor in enumerate(refeicao_exemplo, 1):
    celula = ws_ref.cell(row=4, column=j, value=valor)
    aplicar_estilo_celula(celula, FONTE_NORMAL, PREECHIMENTO_FUNDO, ALINHAMENTO_CENTRO, BORDA_FINA)
    if j in [3, 5]:
        celula.number_format = "dd/mm/yyyy hh:mm"
    elif j == 4:
        celula.number_format = "0.00"

ws_ref.freeze_panes = "A4"

# ============================================================
# FORMATAÇÃO FINAL DAS ABAS
# ============================================================

for ws in [ws_dados, ws_plan, ws_resumo, ws_rec, ws_eventos, ws_ref]:
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

# ============================================================
# SALVAR ARQUIVO
# ============================================================

caminho_arquivo = "/workspaces/aps/APS_PURAN_Base.xlsx"
wb.save(caminho_arquivo)
print(f"Arquivo Excel base criado: {caminho_arquivo}")
