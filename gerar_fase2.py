
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, time, timedelta

# ============================================================
# CONFIGURAÇÕES VISUAIS — SISTEMA APS INDUSTRIAL
# ============================================================

COR_PRIMARIA = "1F4E79"       # Azul escuro
COR_SECUNDARIA = "2E75B6"     # Azul médio
COR_ACENTO = "ED7D31"         # Laranja
COR_SUCESSO = "70AD47"        # Verde
COR_ALERTA = "FF0000"         # Vermelho
COR_AVISO = "FFC000"          # Amarelo
COR_FUNDO = "F2F2F2"          # Cinza claro
COR_BRANCO = "FFFFFF"         # Branco
COR_CINZA_ESCURO = "404040"   # Cinza escuro
COR_CINZA_CLARO = "D9D9D9"    # Cinza claro
COR_VERDE_ESCURO = "375623"   # Verde escuro

FONTE_TITULO = Font(name="Calibri", size=22, bold=True, color=COR_BRANCO)
FONTE_SUBTITULO = Font(name="Calibri", size=14, bold=True, color=COR_BRANCO)
FONTE_CABECALHO = Font(name="Calibri", size=11, bold=True, color=COR_BRANCO)
FONTE_NORMAL = Font(name="Calibri", size=10, color="000000")
FONTE_BOTAO = Font(name="Calibri", size=10, bold=True, color=COR_BRANCO)
FONTE_MENU = Font(name="Calibri", size=11, bold=True, color=COR_BRANCO)
FONTE_KPI = Font(name="Calibri", size=20, bold=True, color=COR_PRIMARIA)
FONTE_ALERTA = Font(name="Calibri", size=10, color=COR_ALERTA)

PREENCHIMENTO_PRIMARIO = PatternFill(start_color=COR_PRIMARIA, end_color=COR_PRIMARIA, fill_type="solid")
PREENCHIMENTO_SECUNDARIO = PatternFill(start_color=COR_SECUNDARIA, end_color=COR_SECUNDARIA, fill_type="solid")
PREENCHIMENTO_ACENTO = PatternFill(start_color=COR_ACENTO, end_color=COR_ACENTO, fill_type="solid")
PREENCHIMENTO_SUCESSO = PatternFill(start_color=COR_SUCESSO, end_color=COR_SUCESSO, fill_type="solid")
PREENCHIMENTO_ALERTA = PatternFill(start_color=COR_ALERTA, end_color=COR_ALERTA, fill_type="solid")
PREENCHIMENTO_AVISO = PatternFill(start_color=COR_AVISO, end_color=COR_AVISO, fill_type="solid")
PREENCHIMENTO_FUNDO = PatternFill(start_color=COR_FUNDO, end_color=COR_FUNDO, fill_type="solid")
PREENCHIMENTO_CINZA_ESCURO = PatternFill(start_color=COR_CINZA_ESCURO, end_color=COR_CINZA_ESCURO, fill_type="solid")
PREENCHIMENTO_CINZA_CLARO = PatternFill(start_color=COR_CINZA_CLARO, end_color=COR_CINZA_CLARO, fill_type="solid")
PREENCHIMENTO_VERDE_ESCURO = PatternFill(start_color=COR_VERDE_ESCURO, end_color=COR_VERDE_ESCURO, fill_type="solid")

BORDA_FINA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC")
)

BORDA_MEDIA = Border(
    left=Side(style="medium", color=COR_SECUNDARIA),
    right=Side(style="medium", color=COR_SECUNDARIA),
    top=Side(style="medium", color=COR_SECUNDARIA),
    bottom=Side(style="medium", color=COR_SECUNDARIA)
)

ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALINHAMENTO_ESQUERDA = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALINHAMENTO_DIREITA = Alignment(horizontal="right", vertical="center")

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def aplicar_estilo_celula(celula, fonte, preenchimento, alinhamento, borda=None):
    celula.font = fonte
    celula.fill = preenchimento
    celula.alignment = alinhamento
    if borda:
        celula.border = borda

def criar_titulo_aba(ws, titulo, subtitulo=""):
    ws.merge_cells("A1:Z1")
    ws["A1"] = titulo
    ws["A1"].font = FONTE_TITULO
    ws["A1"].fill = PREENCHIMENTO_PRIMARIO
    ws["A1"].alignment = ALINHAMENTO_CENTRO
    ws.row_dimensions[1].height = 35

    if subtitulo:
        ws.merge_cells("A2:Z2")
        ws["A2"] = subtitulo
        ws["A2"].font = FONTE_SUBTITULO
        ws["A2"].fill = PREENCHIMENTO_SECUNDARIO
        ws["A2"].alignment = ALINHAMENTO_CENTRO
        ws.row_dimensions[2].height = 22

def criar_cabecalho_tabela(ws, linha, colunas, larguras=None):
    for i, coluna in enumerate(colunas, 1):
        celula = ws.cell(row=linha, column=i, value=coluna)
        aplicar_estilo_celula(celula, FONTE_CABECALHO, PREENCHIMENTO_SECUNDARIO, ALINHAMENTO_CENTRO, BORDA_FINA)
        if larguras and i <= len(larguras):
            ws.column_dimensions[get_column_letter(i)].width = larguras[i-1]

def ajustar_largura_colunas(ws, colunas, largura_padrao=15):
    for i in range(1, colunas + 1):
        letra = get_column_letter(i)
        if ws.column_dimensions[letra].width < largura_padrao:
            ws.column_dimensions[letra].width = largura_padrao

def criar_botao_menu(ws, linha, coluna, texto, macro, largura=140, altura=35):
    celula = ws.cell(row=linha, column=coluna, value=texto)
    aplicar_estilo_celula(celula, FONTE_MENU, PREENCHIMENTO_SECUNDARIO, ALINHAMENTO_CENTRO, BORDA_MEDIA)
    ws.row_dimensions[linha].height = altura / 6
    ws.column_dimensions[get_column_letter(coluna)].width = largura / 8
    return celula

# ============================================================
# CRIAR ARQUIVO
# ============================================================

wb = openpyxl.Workbook()

# Remover aba padrão
wb.remove(wb.active)

# ============================================================
# ABA 1: INICIO (Dashboard Operacional)
# ============================================================

ws_inicio = wb.create_sheet("INICIO")

# Cabeçalho principal
ws_inicio.merge_cells("A1:H1")
ws_inicio["A1"] = "APS PURAN"
ws_inicio["A1"].font = FONTE_TITULO
ws_inicio["A1"].fill = PREENCHIMENTO_PRIMARIO
ws_inicio["A1"].alignment = ALINHAMENTO_CENTRO
ws_inicio.row_dimensions[1].height = 40

ws_inicio.merge_cells("A2:H2")
ws_inicio["A2"] = f"Sistema de Planejamento e Controle da Produção  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws_inicio["A2"].font = Font(name="Calibri", size=11, bold=True, color=COR_BRANCO)
ws_inicio["A2"].fill = PREENCHIMENTO_SECUNDARIO
ws_inicio["A2"].alignment = ALINHAMENTO_CENTRO
ws_inicio.row_dimensions[2].height = 20

# Status do planejamento
ws_inicio.merge_cells("A3:H3")
ws_inicio["A3"] = "STATUS: OPERACIONAL"
ws_inicio["A3"].font = Font(name="Calibri", size=11, bold=True, color=COR_BRANCO)
ws_inicio["A3"].fill = PREENCHIMENTO_SUCESSO
ws_inicio["A3"].alignment = ALINHAMENTO_CENTRO
ws_inicio.row_dimensions[3].height = 22

# Indicadores principais (KPIs)
kpis = [
    ("Total de OPs", "=COUNTA(DADOS!A:A)-1", COR_SECUNDARIA),
    ("Em Produção", '=COUNTIF(DADOS!S:S,"EM PRODUÇÃO")', COR_ACENTO),
    ("Atrasadas", '=COUNTIF(DADOS!S:S,"ATRASADO")', COR_ALERTA),
    ("Concluídas", '=COUNTIF(DADOS!S:S,"CONCLUÍDO")', COR_SUCESSO),
    ("Total Caixas", "=SUM(DADOS!F:F)", COR_PRIMARIA),
    ("Horas Planejadas", "=SUM(DADOS!P:P)", COR_SECUNDARIA),
    ("Horas Atraso", "=SUM(DADOS!T:T)", COR_ALERTA),
    ("Ops c/ Atraso", '=COUNTIF(DADOS!T:T,">0")', COR_AVISO),
]

for i, (titulo, formula, cor) in enumerate(kpis):
    linha = 5 + (i // 4) * 3
    coluna = 1 + (i % 4) * 2
    
    # Título do KPI
    ws_inicio.merge_cells(start_row=linha, start_column=coluna, end_row=linha, end_column=coluna+1)
    celula_titulo = ws_inicio.cell(row=linha, column=coluna, value=titulo)
    celula_titulo.font = FONTE_CABECALHO
    celula_titulo.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
    celula_titulo.alignment = ALINHAMENTO_CENTRO
    celula_titulo.border = BORDA_FINA
    
    # Valor do KPI
    ws_inicio.merge_cells(start_row=linha+1, start_column=coluna, end_row=linha+1, end_column=coluna+1)
    celula_valor = ws_inicio.cell(row=linha+1, column=coluna, value=formula)
    celula_valor.font = FONTE_KPI
    celula_valor.fill = PREENCHIMENTO_FUNDO
    celula_valor.alignment = ALINHAMENTO_CENTRO
    celula_valor.border = BORDA_FINA

# Área de alertas
ws_inicio.merge_cells("A11:H11")
ws_inicio["A11"] = "ALERTAS E NOTIFICAÇÕES"
ws_inicio["A11"].font = FONTE_CABECALHO
ws_inicio["A11"].fill = PREENCHIMENTO_CINZA_ESCURO
ws_inicio["A11"].alignment = ALINHAMENTO_CENTRO
ws_inicio.row_dimensions[11].height = 22

alertas = [
    "• Nenhuma alerta no momento",
    "",
    "• O sistema está operacional",
    "• Clique nos botões abaixo para acessar as funcionalidades"
]

for i, alerta in enumerate(alertas):
    linha = 12 + i
    ws_inicio.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=8)
    celula = ws_inicio.cell(row=linha, column=1, value=alerta)
    celula.font = FONTE_NORMAL
    celula.fill = PREENCHIMENTO_FUNDO
    celula.alignment = ALINHAMENTO_ESQUERDA
    celula.border = BORDA_FINA

# Menu principal
ws_inicio.merge_cells("A18:H18")
ws_inicio["A18"] = "MENU PRINCIPAL"
ws_inicio["A18"].font = FONTE_CABECALHO
ws_inicio["A18"].fill = PREENCHIMENTO_CINZA_ESCURO
ws_inicio["A18"].alignment = ALINHAMENTO_CENTRO
ws_inicio.row_dimensions[18].height = 22

# Botões do menu
menu_botoes = [
    ("📋 Ordens de Produção", "IrParaOPs", 20, 40),
    ("🏭 Máquinas", "IrParaMaquinas", 20, 40),
    ("📅 Planejamento", "IrParaPlanejamento", 20, 40),
    ("⚠ Atrasos", "AtualizarAtrasosAPS", 20, 40),
    ("🔧 Eventos/Manutenção", "AplicarEventosAPS", 20, 40),
    ("🍽 Refeições", "DesenharRefeicoes", 20, 40),
    ("🔄 Recalcular APS", "AtualizarAPS", 20, 40),
    ("⚙ Configurações", "IrParaConfig", 20, 40),
]

for i, (texto, macro, largura, altura) in enumerate(menu_botoes):
    linha = 20 + (i // 4) * 2
    coluna = 1 + (i % 4) * 2
    
    ws_inicio.merge_cells(start_row=linha, start_column=coluna, end_row=linha, end_column=coluna+1)
    celula = ws_inicio.cell(row=linha, column=coluna, value=texto)
    aplicar_estilo_celula(celula, FONTE_MENU, PREENCHIMENTO_SECUNDARIO, ALINHAMENTO_CENTRO, BORDA_MEDIA)
    ws_inicio.row_dimensions[linha].height = altura / 6

# Formatação final
ws_inicio.sheet_view.showGridLines = False
ws_inicio.sheet_view.zoomScale = 90

# ============================================================
# ABA 2: OPs (Ordens de Produção)
# ============================================================

ws_ops = wb.create_sheet("OPs")

criar_titulo_aba(ws_ops, "ORDENS DE PRODUÇÃO", "Cadastro e consulta de OPs")

# Área de busca
ws_ops.merge_cells("A4:F4")
ws_ops["A4"] = "BUSCAR OP"
ws_ops["A4"].font = FONTE_CABECALHO
ws_ops["A4"].fill = PREENCHIMENTO_CINZA_CLARO
ws_ops["A4"].alignment = ALINHAMENTO_ESQUERDA
ws_ops.row_dimensions[4].height = 20

# Filtros
filtros = ["OP:", "Produto:", "Máquina:", "Status:"]
for i, filtro in enumerate(filtros):
    coluna = 1 + i * 2
    ws_ops.cell(row=5, column=coluna, value=filtro).font = FONTE_NORMAL
    ws_ops.cell(row=5, column=coluna+1).fill = PREENCHIMENTO_FUNDO
    ws_ops.cell(row=5, column=coluna+1).border = BORDA_FINA
    ws_ops.row_dimensions[5].height = 22

# Botões de ação
ws_ops.merge_cells("A7:H7")
ws_ops["A7"] = "AÇÕES"
ws_ops["A7"].font = FONTE_CABECALHO
ws_ops["A7"].fill = PREENCHIMENTO_CINZA_CLARO
ws_ops["A7"].alignment = ALINHAMENTO_ESQUERDA

acoes = [
    ("+ Nova OP", "AbrirCadastroOP", 18),
    ("📊 Ver Planejamento", "IrParaPlanejamento", 18),
    ("🔄 Recalcular", "AtualizarAPS", 18),
]

for i, (texto, macro, largura) in enumerate(acoes):
    coluna = 1 + i * 3
    ws_ops.merge_cells(start_row=8, start_column=coluna, end_row=8, end_column=coluna+2)
    celula = ws_ops.cell(row=8, column=coluna, value=texto)
    aplicar_estilo_celula(celula, FONTE_BOTAO, PREENCHIMENTO_SECUNDARIO, ALINHAMENTO_CENTRO, BORDA_MEDIA)

# Tabela de OPs
colunas_ops = [
    "OP", "Produto", "Dosagem", "Quantidade", "Unidade",
    "Caixas", "Máquina", "Sequência", "Data Planejada",
    "Velocidade", "OEE", "Capacidade_h", "Produção_h", "Setup_h",
    "Duração Base_h", "Duração Total (h)", "Início", "Fim",
    "Status", "Atraso_h", "Eventos_h", "Refeição_h", "Início Original", "Fim Original", "Observação"
]

larguras_ops = [12, 20, 12, 12, 10, 10, 15, 10, 16, 10, 8, 12, 12, 10, 14, 16, 16, 16, 14, 10, 10, 12, 16, 16, 25]

criar_cabecalho_tabela(ws_ops, 10, colunas_ops, larguras_ops)

# Dados de exemplo
dados_exemplo = [
    ["OP001", "PURAN", "25 mcg", 1000, "COMP", 200, "FETTE 2090", 1, datetime(2026, 8, 25, 5, 0), 500, 0.85, 425, 0.47, 0.5, 0.97, 1.0, datetime(2026, 8, 25, 5, 0), datetime(2026, 8, 25, 6, 0), "PLANEJADO", 0, 0, 0, "", "", ""],
    ["OP002", "PURAN", "50 mcg", 800, "COMP", 160, "FETTE 2", 1, datetime(2026, 8, 25, 6, 30), 450, 0.80, 360, 0.44, 0.5, 1.33, 1.33, datetime(2026, 8, 25, 6, 30), datetime(2026, 8, 25, 8, 0), "PLANEJADO", 0, 0, 0, "", "", ""],
    ["OP003", "DIPIRONA", "1g", 500, "CX", 500, "MEDISEAL", 1, datetime(2026, 8, 25, 5, 0), 600, 0.75, 450, 1.11, 0.5, 1.61, 1.61, datetime(2026, 8, 25, 5, 0), datetime(2026, 8, 25, 6, 45), "PLANEJADO", 0, 0, 0, "", "", ""],
]

for i, linha_dados in enumerate(dados_exemplo, 11):
    for j, valor in enumerate(linha_dados, 1):
        celula = ws_ops.cell(row=i, column=j, value=valor)
        aplicar_estilo_celula(celula, FONTE_NORMAL, PREENCHIMENTO_FUNDO, ALINHAMENTO_CENTRO if j != 25 else ALINHAMENTO_ESQUERDA, BORDA_FINA)
        if j in [16, 17, 18]:
            celula.number_format = "dd/mm/yyyy hh:mm"
        elif j in [10, 11, 12, 13, 14, 15, 20, 21]:
            celula.number_format = "0.00"

ws_ops.freeze_panes = "A11"

# ============================================================
# ABA 3: MAQUINAS (Cadastro de Recursos)
# ============================================================

ws_maq = wb.create_sheet("MAQUINAS")

criar_titulo_aba(ws_maq, "CADASTRO DE MÁQUINAS", "Gerenciamento de recursos produtivos")

# Botão nova máquina
ws_maq.merge_cells("A4:F4")
ws_maq["A4"] = "+ Nova Máquina"
ws_maq["A4"].font = FONTE_BOTAO
ws_maq["A4"].fill = PREENCHIMENTO_SUCESSO
ws_maq["A4"].alignment = ALINHAMENTO_CENTRO
ws_maq["A4"].border = BORDA_MEDIA
ws_maq.row_dimensions[4].height = 25

# Tabela de máquinas
colunas_maq = ["Máquina", "Velocidade", "Unidade Vel.", "OEE (%)", "Setup (h)", "CpC", "Ativa", "Observação"]
larguras_maq = [18, 12, 12, 12, 12, 12, 10, 30]

criar_cabecalho_tabela(ws_maq, 6, colunas_maq, larguras_maq)

maquinas_exemplo = [
    ["FETTE 2090", 500, "cx/h", 0.85, 0.5, 20, "Sim", "Envelope"],
    ["FETTE 2", 450, "cx/h", 0.80, 0.5, 20, "Sim", "Envelope"],
    ["MEDISEAL", 600, "cx/h", 0.75, 0.5, 20, "Sim", "Blister"],
    ["BLISTERFLEX", 550, "cx/h", 0.80, 0.5, 20, "Sim", "Blister"],
]

for i, linha_dados in enumerate(maquinas_exemplo, 7):
    for j, valor in enumerate(linha_dados, 1):
        celula = ws_maq.cell(row=i, column=j, value=valor)
        aplicar_estilo_celula(celula, FONTE_NORMAL, PREENCHIMENTO_FUNDO, ALINHAMENTO_CENTRO, BORDA_FINA)
        if j == 4:  # OEE
            celula.number_format = "0%"
        elif j in [2, 3, 5, 6]:
            celula.number_format = "0.00"

ws_maq.freeze_panes = "A7"

# ============================================================
# ABA 4: PLANEJAMENTO (Timeline/Gantt)
# ============================================================

ws_plan = wb.create_sheet("PLANEJAMENTO")

criar_titulo_aba(ws_plan, "PLANEJAMENTO DA PRODUÇÃO", "Timeline visual e cards das OPs")

# Botões de ação
ws_plan.merge_cells("A4:P4")
ws_plan["A4"] = "🔄 Recalcular   ⚠ Atrasos   🔧 Eventos   🍽 Refeições   🏠 Dashboard"
ws_plan["A4"].font = FONTE_BOTAO
ws_plan["A4"].fill = PREENCHIMENTO_SECUNDARIO
ws_plan["A4"].alignment = ALINHAMENTO_CENTRO
ws_plan["A4"].border = BORDA_MEDIA
ws_plan.row_dimensions[4].height = 25

# Cabeçalho da timeline
ws_plan.cell(row=6, column=1, value="MÁQUINA").font = FONTE_CABECALHO
ws_plan.cell(row=6, column=1).fill = PREENCHIMENTO_CINZA_ESCURO
ws_plan.cell(row=6, column=1).alignment = ALINHAMENTO_CENTRO
ws_plan.cell(row=6, column=1).border = BORDA_FINA

hora_inicio = time(5, 0)
for i in range(2, 50):
    hora = (datetime.combine(datetime.today(), hora_inicio) + timedelta(minutes=30 * (i - 2))).time()
    celula = ws_plan.cell(row=6, column=i, value=hora.strftime("%H:%M"))
    celula.font = FONTE_CABECALHO
    celula.fill = PREENCHIMENTO_CINZA_ESCURO
    celula.alignment = ALINHAMENTO_CENTRO
    celula.border = BORDA_FINA
    ws_plan.column_dimensions[get_column_letter(i)].width = 9

# Máquinas
maquinas = ["FETTE 2090", "FETTE 2", "MEDISEAL", "BLISTERFLEX"]
for i, maquina in enumerate(maquinas, 7):
    celula = ws_plan.cell(row=i, column=1, value=maquina)
    celula.font = Font(name="Calibri", size=10, bold=True, color=COR_BRANCO)
    celula.fill = PREENCHIMENTO_SECUNDARIO
    celula.alignment = ALINHAMENTO_CENTRO
    celula.border = BORDA_FINA
    ws_plan.row_dimensions[i].height = 60

ws_plan.column_dimensions["A"].width = 20
ws_plan.freeze_panes = "B7"

# ============================================================
# ABA 5: CONFIG (Oculta — Configurações)
# ============================================================

ws_config = wb.create_sheet("CONFIG")
ws_config.sheet_state = "hidden"

# ============================================================
# ABA 6: DADOS (Oculta — Backend)
# ============================================================

ws_dados = wb.create_sheet("DADOS")
ws_dados.sheet_state = "hidden"

colunas_dados = [
    "OP", "Produto", "Dosagem", "Quantidade", "Unidade",
    "Caixas", "Máquina", "Sequência", "Data Planejada",
    "Velocidade", "OEE", "Capacidade_h", "Produção_h", "Setup_h",
    "Duração Base_h", "Duração Total (h)", "Início", "Fim",
    "Status", "Atraso_h", "Eventos_h", "Refeição_h", "Início Original", "Fim Original", "Observação"
]

for i, coluna in enumerate(colunas_dados, 1):
    ws_dados.cell(row=1, column=i, value=coluna)

ws_dados.cell(row=1, column=1).font = FONTE_CABECALHO

# ============================================================
# ABA 7: RECURSOS (Oculta — Backend)
# ============================================================

ws_rec = wb.create_sheet("RECURSOS")
ws_rec.sheet_state = "hidden"

colunas_rec = ["Máquina", "Velocidade", "Unidade Velocidade", "OEE", "Comprimidos por caixa", "Setup padrão", "Disponibilidade", "Observação"]
for i, coluna in enumerate(colunas_rec, 1):
    ws_rec.cell(row=1, column=i, value=coluna)

# ============================================================
# ABA 8: EVENTOS (Oculta — Backend)
# ============================================================

ws_eventos = wb.create_sheet("EVENTOS")
ws_eventos.sheet_state = "hidden"

colunas_eventos = ["OP", "Tipo", "Duração_h", "Ativo", "Aplicado"]
for i, coluna in enumerate(colunas_eventos, 1):
    ws_eventos.cell(row=1, column=i, value=coluna)

# ============================================================
# ABA 9: REFEICAO (Oculta — Backend)
# ============================================================

ws_ref = wb.create_sheet("REFEICAO")
ws_ref.sheet_state = "hidden"

colunas_ref = ["ID", "Máquina", "Início", "Duração_h", "Fim", "Ativo", "Observação"]
for i, coluna in enumerate(colunas_ref, 1):
    ws_ref.cell(row=1, column=i, value=coluna)

# ============================================================
# ABA 10: RESUMO (Oculta — Backend Dashboard)
# ============================================================

ws_resumo = wb.create_sheet("RESUMO")
ws_resumo.sheet_state = "hidden"

# ============================================================
# FORMATAÇÃO FINAL DAS ABAS VISÍVEIS
# ============================================================

for ws in [ws_inicio, ws_ops, ws_maq, ws_plan]:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

# ============================================================
# SALVAR ARQUIVO
# ============================================================

caminho_arquivo = "/workspaces/aps/APS_PURAN_Fase2.xlsx"
wb.save(caminho_arquivo)
print(f"Arquivo Fase 2 criado: {caminho_arquivo}")
